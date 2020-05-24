# api.py
from flask import Flask, request, render_template
from preprocessing import preprocessing
from werkzeug.utils import secure_filename 
from openpyxl import load_workbook
import pickle
import mysql.connector
import os
import json
import simplejson


db = mysql.connector.connect(
  host="localhost",
  user="root",
  passwd="",
  database = "skripsi"
)

app = Flask(__name__)
vectorizer = pickle.load(open("vectorizer.b", "rb"))
ex = pickle.load(open("ex.b", "rb"))


def parseSentiment(sentiment):
  if sentiment == 1.0:
    return "1"
  else:
    return "0"
# return "0"

# localhost:5000/
# @app.route("/")
# def homepage():
#     html = "<h1>Hello</h1>"
#     html += "Prediksi sentimen"
#     return html


# localhost:5000/help

@app.route('/')
def index():
    return render_template('index.html')

@app.route("/help")
def help():
    html = "<h1>Help</h1>"
    return html


@app.route("/predict")
def predict():
    input_user = request.args.get("input", "")
    preproces = preprocessing(str(input_user))
    X = vectorizer.transform([preproces])
    sentiment = ex.predict(X)
    

    cursor = db.cursor()
    # sql = "INSERT INTO tanggapan (pesan, sentimen) VALUES (%s, %s)"
    # val = (input_user, parseSentiment(sentiment[0]))
    # cursor.execute(sql, val)

    db.commit()
    
    html = "<h1>Predict</h1>"
    html += "Input: " + input_user
    html += "<br>"
    html += "Hasil preprocessing: " + preproces
    html += "<br>"
    html += "Hasil prediksi: " + parseSentiment(sentiment[0])
    
    print("{} data ditambahkan".format(cursor.rowcount))
    return html


# contoh cara menangkap post request
# parameter input name 'test'
@app.route("/test", methods=["POST"])
def test():
    return request.form["test"]

ALLOWED_EXTENSION = set(['xlsx'])
app.config['UPLOAD_FOLDER'] = 'uploads'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSION
   
@app.route('/input', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
 
        file = request.files['file']
 
        if 'file' not in request.files:
            # return render_template('input.html')
            return  'file tidak dapat di simpan' + ' <a href="/input">kembali</a>'
 
        if file.filename == '':
            return render_template('input.html')
 
        if file and allowed_file(file.filename):
            # filename = secure_filename(file.filename)
            filename = 'data'+'.xlsx'
            file.save(os.path.join(app.config['UPLOAD_FOLDER'],  filename))            
            # return  'file ' + filename +' di simpan' + ' <a href="/uploads">kembali</a>'
 
    return render_template('input.html')


@app.route('/proses')
def proses():

    dat_file    = "uploads/data.xlsx"
    wb          = load_workbook(dat_file)    
    sheet       = wb.active
    jum_baris   = sheet.max_row
    # jum_baris   = 15

    for i in range(2, jum_baris):
        waktu    = sheet.cell(row=i,column=1)
        nim    = sheet.cell(row=i,column=3)
        pelatihan    = sheet.cell(row=i,column=5)
        pesan   = sheet.cell(row=i,column=9)

        preproces = preprocessing(str(pesan))
        X = vectorizer.transform([preproces])
        sentiment = ex.predict(X)
        senti = parseSentiment(sentiment[0])
        # print(i, "=", sentiment[0],"=",pesan.value," | ",senti)
        # senti = None
        # senti = parseSentiment(sentiment[0])

        cursor = db.cursor()
        sql = "INSERT INTO tanggapan (waktu, nim, pelatihan, pesan, sentimen) VALUES (%s, %s, %s, %s, %s)"
        val = (waktu.value, nim.value, pelatihan.value, pesan.value, senti)
        # preproces   = None
        # X           = None
        # sentiment   = None
        # senti       = None
        cursor.execute(sql, val)

        db.commit()
        # html = "<h1>Predict</h1>"
        # html += "Input: " + pesan.value
        # html += "<br>"
        # html += "Hasil preprocessing: " + preproces
        # html += "<br>"
        
        # print("{} data ditambahkan".format(cursor.rowcount))
    # return "0"
    return render_template('rekap.html')

@app.route('/rekap')
def rekap():

        cursor = db.cursor()
        # sql = "SELECT waktu, pelatihan, sentimen FROM tanggapan"
        # sql = "SELECT waktu, pelatihan, COUNT(sentimen) as sentimen FROM tanggapan WHERE sentimen=1 GROUP BY waktu"
        sql = "SELECT pesan_id, waktu, pesan_asli, replace(replace(sentimen, '1', 'positif'), '0', 'negatif') FROM tanggapan GROUP BY waktu, sentimen"
        # val = (pesan_id.value, waktu.value, nim.value, pelatihan.value, pesan.value, 0)
        cursor.execute(sql)
        rv = cursor.fetchall()
        return render_template("rekap.html", value=rv)

# @app.route('proses')
# def proses():

#   # Open the workbook and define the worksheet
#   book = xlrd.open_workbook("files/data.xlsx")
#   sheet = book.sheet_by_index(1)

#   return html


@app.route('/grafik')
def grafik():
    
    return render_template('grafik.php')

# @app.route('/graph')
# def graph():

#     pcen = db.cursor()
#     sql_pcen = "SELECT round((SELECT COUNT(*) FROM `tanggapan` WHERE sentimen = '1') / (SELECT COUNT(*) FROM `tanggapan`) * 100) as positif, round((SELECT COUNT(*) FROM `tanggapan` WHERE sentimen = '0') / (SELECT COUNT(*) FROM `tanggapan`) * 100) as negatif from tanggapan LIMIT 1"
#     pcen.execute(sql_pcen)
#     pce = pcen.fetchall()

#     arr=[]
#     for product in pce:
#         vals = {}
#         vals['Positif']=product[0]
#         vals['Negatif']=product[1]
#         arr.append(vals)
#     jsongr = json.dumps(arr)

#     return jsongr

if __name__ == "__main__":
    app.run(debug=True, port=5000)
